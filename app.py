from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    send_file,
)

import os
import io
from datetime import datetime

from werkzeug.utils import secure_filename

from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import or_

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = "change-this-later"

# Upload folder
UPLOAD_FOLDER = os.path.join(app.root_path, "static", "uploads")
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Allowed file types
ALLOWED_EXTENSIONS = {"pdf", "png", "jpg", "jpeg"}

# Database
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///members.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)


def allowed_file(filename):
    return (
        "." in filename and
        filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS
    )


# ==========================
# Member Model
# ==========================

class Member(db.Model):
    __tablename__ = "members"

    member_id = db.Column(db.Integer, primary_key=True)

    full_name = db.Column(
        db.String(100),
        nullable=False
    )

    id_number = db.Column(
        db.String(20),
        unique=True,
        nullable=False
    )

    next_of_kin = db.Column(
        db.String(100),
        nullable=False
    )

    id_upload = db.Column(
        db.String(255),
        nullable=False
    )

    registration_date = db.Column(
        db.DateTime,
        default=datetime.utcnow
    )


with app.app_context():
    db.create_all()


# ==========================
# Dashboard
# ==========================

@app.route("/")
def dashboard():

    total_members = Member.query.count()

    return render_template(
        "dashboard.html",
        total_members=total_members
    )


# ==========================
# Register Member
# ==========================

@app.route("/register", methods=["GET", "POST"])
def register():

    if request.method == "POST":

        full_name = request.form["full_name"]
        id_number = request.form["id_number"]
        next_of_kin = request.form["next_of_kin"]

        uploaded_file = request.files["id_upload"]

        if uploaded_file.filename == "":
            flash("Please select an ID document.", "danger")
            return redirect(url_for("register"))

        if not allowed_file(uploaded_file.filename):
            flash(
                "Only PDF, JPG, JPEG and PNG files are allowed.",
                "danger"
            )
            return redirect(url_for("register"))

        existing_member = Member.query.filter_by(
            id_number=id_number
        ).first()

        if existing_member:
            flash(
                "A member with this ID number already exists.",
                "danger"
            )
            return redirect(url_for("register"))

        filename = secure_filename(uploaded_file.filename)

        uploaded_file.save(
            os.path.join(app.config["UPLOAD_FOLDER"], filename)
        )

        member = Member(
            full_name=full_name,
            id_number=id_number,
            next_of_kin=next_of_kin,
            id_upload=filename
        )

        db.session.add(member)
        db.session.commit()

        flash(
            "Member registered successfully!",
            "success"
        )

        return redirect(url_for("members"))

    return render_template("register_member.html")
# ==========================
# View Members
# ==========================

@app.route("/members")
def members():

    search = request.args.get("search", "")
    page = request.args.get("page", 1, type=int)

    query = Member.query

    if search:

        query = query.filter(
            or_(
                Member.full_name.ilike(f"%{search}%"),
                Member.id_number.ilike(f"%{search}%")
            )
        )

    members = query.order_by(
        Member.member_id.desc()
    ).paginate(
        page=page,
        per_page=5,
        error_out=False
    )

    return render_template(
        "members.html",
        members=members,
        search=search
    )


# ==========================
# View Single Member
# ==========================

@app.route("/member/<int:member_id>")
def view_member(member_id):

    member = Member.query.get_or_404(member_id)

    return render_template(
        "view_member.html",
        member=member
    )


# ==========================
# Edit Member
# ==========================

@app.route("/edit/<int:member_id>", methods=["GET", "POST"])
def edit_member(member_id):

    member = Member.query.get_or_404(member_id)

    if request.method == "POST":

        existing_member = Member.query.filter(
            Member.id_number == request.form["id_number"],
            Member.member_id != member_id
        ).first()

        if existing_member:

            flash(
                "A member with this ID number already exists.",
                "danger"
            )

            return redirect(
                url_for(
                    "edit_member",
                    member_id=member_id
                )
            )

        member.full_name = request.form["full_name"]
        member.id_number = request.form["id_number"]
        member.next_of_kin = request.form["next_of_kin"]

        db.session.commit()

        flash(
            "Member updated successfully!",
            "success"
        )

        return redirect(url_for("members"))

    return render_template(
        "edit_member.html",
        member=member
    )
# ==========================
# Delete Member
# ==========================

@app.route("/delete/<int:member_id>")
def delete_member(member_id):

    member = Member.query.get_or_404(member_id)

    file_path = os.path.join(
        app.config["UPLOAD_FOLDER"],
        member.id_upload
    )

    if os.path.exists(file_path):
        os.remove(file_path)

    db.session.delete(member)
    db.session.commit()

    flash(
        "Member deleted successfully!",
        "success"
    )

    return redirect(url_for("members"))


# ==========================
# Export Members to Excel
# ==========================

@app.route("/export")
def export_members():

    members = Member.query.order_by(Member.member_id.asc()).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Members"

    # Report Title
    sheet["A1"] = "MEMBER MANAGEMENT SYSTEM"
    sheet["A1"].font = Font(bold=True, size=16)

    # Report Date
    sheet["A2"] = (
        f"Generated on: "
        f"{datetime.now().strftime('%d %B %Y %H:%M')}"
    )

    # Blank Row
    sheet.append([])

    # Headers
    headers = [
        "Member ID",
        "Full Name",
        "National ID",
        "Next of Kin",
        "Registration Date"
    ]

    sheet.append(headers)

    # Make headers bold
    for cell in sheet[4]:
        cell.font = Font(bold=True)

    # Member Data
    for member in members:

        sheet.append([
            member.member_id,
            member.full_name,
            member.id_number,
            member.next_of_kin,
            member.registration_date.strftime("%d-%m-%Y")
            if member.registration_date else ""
        ])

    # Auto-size columns
    for column in sheet.columns:

        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:

            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        sheet.column_dimensions[column_letter].width = max_length + 3

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name="Members.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ==========================
# Run Application
# ==========================

if __name__ == "__main__":
    app.run(debug=True)